from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction
from django.db.models import Sum, Q
from django.db.models.functions import Coalesce
from decimal import Decimal
from ..models import (
    Bodegas, Lotes, MovimientosInventario, 
    Materiales, Maquinas, Operarios, Terceros,
    ProduccionMolido, ProduccionLavado, ProduccionPeletizado, ProduccionInyeccion,
    ResiduosProduccion, ProduccionConsumo, MotivoParo, ParosProduccion,
    Despacho, DetalleDespacho
)
from ..inventario_utils import procesar_movimiento_inventario
from django.core.exceptions import ValidationError
from datetime import datetime, timedelta
import uuid
from django.utils import timezone
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_GET
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def lotes(request):
    """Vista para gestionar los lotes."""
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Obtener datos del formulario
                numero_lote = request.POST.get('numero_lote')
                id_material = request.POST.get('id_material')
                cantidad_inicial = Decimal(request.POST.get('cantidad_inicial'))
                unidad_medida = request.POST.get('unidad_medida')
                id_bodega = request.POST.get('id_bodega')
                fecha_vencimiento = request.POST.get('fecha_vencimiento')
                costo_unitario = request.POST.get('costo_unitario')
                proveedor = request.POST.get('proveedor')
                clasificacion = request.POST.get('clasificacion')
                observaciones = request.POST.get('observaciones')

                # Crear nuevo lote
                Lotes.objects.create(
                    numero_lote=numero_lote,
                    id_material_id=id_material,
                    cantidad_inicial=cantidad_inicial,
                    cantidad_actual=cantidad_inicial,
                    unidad_medida=unidad_medida,
                    id_bodega_actual_id=id_bodega,
                    fecha_vencimiento=fecha_vencimiento if fecha_vencimiento else None,
                    costo_unitario=costo_unitario if costo_unitario else None,
                    proveedor_origen_id=proveedor if proveedor else None,
                    clasificacion=clasificacion if clasificacion else None,
                    observaciones=observaciones,
                    activo=True
                )
                messages.success(request, 'Lote creado exitosamente.')
                return redirect('gestion:lotes')
        except Exception as e:
            messages.error(request, f'Error al crear el lote: {str(e)}')

    lotes_list = Lotes.objects.all().order_by('-fecha_creacion')
    materiales_list = Materiales.objects.all()
    bodegas_list = Bodegas.objects.all()
    proveedores_list = Terceros.objects.filter(tipo='Proveedor')
    
    context = {
        'lotes': lotes_list,
        'materiales': materiales_list,
        'bodegas': bodegas_list,
        'proveedores': proveedores_list,
    }
    return render(request, 'gestion/lotes.html', context)

@login_required
def maquinas(request):
    """Vista para gestionar las máquinas."""
    if request.method == 'POST':
        try:
            nombre = request.POST.get('nombre')
            tipo_proceso = request.POST.get('tipo_proceso')
            descripcion = request.POST.get('descripcion')
            activo = request.POST.get('activo', 'off') == 'on'

            Maquinas.objects.create(
                nombre=nombre,
                tipo_proceso=tipo_proceso,
                descripcion=descripcion,
                activo=activo
            )
            messages.success(request, 'Máquina creada exitosamente.')
            return redirect('gestion:maquinas')
        except Exception as e:
            messages.error(request, f'Error al crear la máquina: {str(e)}')

    maquinas_list = Maquinas.objects.all()
    context = {
        'maquinas': maquinas_list,
    }
    return render(request, 'gestion/maquinas.html', context)

@login_required
def materiales(request):
    """Vista para gestionar los materiales."""
    if request.method == 'POST':
        try:
            nombre = request.POST.get('nombre')
            tipo = request.POST.get('tipo')
            descripcion = request.POST.get('descripcion', '')

            Materiales.objects.create(
                nombre=nombre,
                tipo=tipo,
                descripcion=descripcion
            )
            messages.success(request, 'Material creado exitosamente.')
            return redirect('gestion:materiales')
        except Exception as e:
            messages.error(request, f'Error al crear el material: {str(e)}')

    materiales_list = Materiales.objects.all()
    context = {
        'materiales': materiales_list,
    }
    return render(request, 'gestion/materiales.html', context)

@login_required
def operarios(request):
    """Vista para gestionar los operarios."""
    if request.method == 'POST':
        try:
            codigo = request.POST.get('codigo')
            nombre_completo = request.POST.get('nombre_completo')
            activo = request.POST.get('activo', 'off') == 'on'

            Operarios.objects.create(
                codigo=codigo,
                nombre_completo=nombre_completo,
                activo=activo
            )
            messages.success(request, 'Operario creado exitosamente.')
            return redirect('gestion:operarios')
        except Exception as e:
            messages.error(request, f'Error al crear el operario: {str(e)}')

    operarios_list = Operarios.objects.all()
    context = {
        'operarios': operarios_list,
    }
    return render(request, 'gestion/operarios.html', context)

@login_required
def terceros(request):
    """Vista para gestionar los terceros."""
    if request.method == 'POST':
        try:
            nombre = request.POST.get('nombre')
            tipo = request.POST.get('tipo')
            identificacion = request.POST.get('identificacion')
            direccion = request.POST.get('direccion')
            telefono = request.POST.get('telefono')
            email = request.POST.get('email')

            Terceros.objects.create(
                nombre=nombre,
                tipo=tipo,
                identificacion=identificacion,
                direccion=direccion,
                telefono=telefono,
                email=email
            )
            messages.success(request, 'Tercero creado exitosamente.')
            return redirect('gestion:terceros')
        except Exception as e:
            messages.error(request, f'Error al crear el tercero: {str(e)}')

    terceros_list = Terceros.objects.all()
    context = {
        'terceros': terceros_list,
    }
    return render(request, 'gestion/terceros.html', context)

@login_required
def bodegas(request):
    """Vista para gestionar las bodegas."""
    if request.method == 'POST':
        try:
            nombre = request.POST.get('nombre')
            descripcion = request.POST.get('descripcion', '')

            Bodegas.objects.create(
                nombre=nombre,
                descripcion=descripcion
            )
            messages.success(request, 'Bodega creada exitosamente.')
            return redirect('gestion:bodegas')
        except Exception as e:
            messages.error(request, f'Error al crear la bodega: {str(e)}')

    bodegas_list = Bodegas.objects.all()
    context = {
        'bodegas': bodegas_list,
    }
    return render(request, 'gestion/bodegas.html', context)

# Funciones de edición
@login_required
@transaction.atomic
def editar_lote(request, id):
    """Vista para editar un lote."""
    lote = get_object_or_404(Lotes, pk=id)
    if request.method == 'POST':
        try:
            lote.numero_lote = request.POST.get('numero_lote')
            lote.id_material_id = request.POST.get('id_material')
            lote.id_bodega_actual_id = request.POST.get('id_bodega')
            lote.fecha_vencimiento = request.POST.get('fecha_vencimiento') or None
            lote.costo_unitario = request.POST.get('costo_unitario') or None
            lote.proveedor_origen_id = request.POST.get('proveedor') or None
            lote.clasificacion = request.POST.get('clasificacion') or None
            lote.observaciones = request.POST.get('observaciones')
            lote.save()
            messages.success(request, 'Lote actualizado exitosamente.')
            return redirect('gestion:lotes')
        except Exception as e:
            messages.error(request, f'Error al actualizar el lote: {str(e)}')
    
    context = {
        'lote': lote,
        'materiales': Materiales.objects.all(),
        'bodegas': Bodegas.objects.all(),
        'proveedores': Terceros.objects.filter(tipo='Proveedor'),
    }
    return render(request, 'gestion/editar_lote.html', context)

@login_required
@transaction.atomic
def editar_maquina(request, id):
    """Vista para editar una máquina."""
    maquina = get_object_or_404(Maquinas, pk=id)
    if request.method == 'POST':
        try:
            maquina.nombre = request.POST.get('nombre')
            maquina.tipo_proceso = request.POST.get('tipo_proceso')
            maquina.descripcion = request.POST.get('descripcion')
            maquina.activo = request.POST.get('activo', 'off') == 'on'
            maquina.save()
            messages.success(request, 'Máquina actualizada exitosamente.')
            return redirect('gestion:maquinas')
        except Exception as e:
            messages.error(request, f'Error al actualizar la máquina: {str(e)}')
    
    context = {
        'maquina': maquina,
    }
    return render(request, 'gestion/editar_maquina.html', context)

@login_required
@transaction.atomic
def editar_material(request, id):
    """Vista para editar un material."""
    material = get_object_or_404(Materiales, pk=id)
    if request.method == 'POST':
        try:
            material.nombre = request.POST.get('nombre')
            material.tipo = request.POST.get('tipo')
            material.descripcion = request.POST.get('descripcion')
            material.save()
            messages.success(request, 'Material actualizado exitosamente.')
            return redirect('gestion:materiales')
        except Exception as e:
            messages.error(request, f'Error al actualizar el material: {str(e)}')
    
    context = {
        'material': material,
    }
    return render(request, 'gestion/editar_material.html', context)

@login_required
@transaction.atomic
def editar_operario(request, id):
    """Vista para editar un operario."""
    operario = get_object_or_404(Operarios, pk=id)
    if request.method == 'POST':
        try:
            operario.codigo = request.POST.get('codigo')
            operario.nombre_completo = request.POST.get('nombre_completo')
            operario.activo = request.POST.get('activo', 'off') == 'on'
            operario.save()
            messages.success(request, 'Operario actualizado exitosamente.')
            return redirect('gestion:operarios')
        except Exception as e:
            messages.error(request, f'Error al actualizar el operario: {str(e)}')
    
    context = {
        'operario': operario,
    }
    return render(request, 'gestion/editar_operario.html', context)

@login_required
@transaction.atomic
def editar_tercero(request, id):
    """Vista para editar un tercero."""
    tercero = get_object_or_404(Terceros, pk=id)
    if request.method == 'POST':
        try:
            tercero.nombre = request.POST.get('nombre')
            tercero.tipo = request.POST.get('tipo')
            tercero.identificacion = request.POST.get('identificacion')
            tercero.direccion = request.POST.get('direccion')
            tercero.telefono = request.POST.get('telefono')
            tercero.email = request.POST.get('email')
            tercero.save()
            messages.success(request, 'Tercero actualizado exitosamente.')
            return redirect('gestion:terceros')
        except Exception as e:
            messages.error(request, f'Error al actualizar el tercero: {str(e)}')
    
    context = {
        'tercero': tercero,
    }
    return render(request, 'gestion/editar_tercero.html', context)

@login_required
@transaction.atomic
def editar_bodega(request, id):
    """Vista para editar una bodega."""
    bodega = get_object_or_404(Bodegas, pk=id)
    if request.method == 'POST':
        try:
            bodega.nombre = request.POST.get('nombre')
            bodega.descripcion = request.POST.get('descripcion', '')
            bodega.save()
            messages.success(request, 'Bodega actualizada exitosamente.')
            return redirect('gestion:bodegas')
        except Exception as e:
            messages.error(request, f'Error al actualizar la bodega: {str(e)}')
    
    context = {
        'bodega': bodega,
    }
    return render(request, 'gestion/editar_bodega.html', context)

# Funciones de eliminación
@login_required
@transaction.atomic
def eliminar_lote(request, id):
    """Vista para eliminar un lote."""
    lote = get_object_or_404(Lotes, pk=id)
    if request.method == 'POST':
        try:
            lote.activo = False
            lote.save()
            messages.success(request, 'Lote eliminado exitosamente.')
            return redirect('gestion:lotes')
        except Exception as e:
            messages.error(request, f'Error al eliminar el lote: {str(e)}')
    return redirect('gestion:lotes')

@login_required
@transaction.atomic
def eliminar_maquina(request, id):
    """Vista para eliminar una máquina."""
    maquina = get_object_or_404(Maquinas, pk=id)
    if request.method == 'POST':
        try:
            maquina.delete()
            messages.success(request, 'Máquina eliminada exitosamente.')
            return redirect('gestion:maquinas')
        except Exception as e:
            messages.error(request, f'Error al eliminar la máquina: {str(e)}')
    return redirect('gestion:maquinas')

@login_required
@transaction.atomic
def eliminar_material(request, id):
    """Vista para eliminar un material."""
    material = get_object_or_404(Materiales, pk=id)
    if request.method == 'POST':
        try:
            material.delete()
            messages.success(request, 'Material eliminado exitosamente.')
            return redirect('gestion:materiales')
        except Exception as e:
            messages.error(request, f'Error al eliminar el material: {str(e)}')
    return redirect('gestion:materiales')

@login_required
@transaction.atomic
def eliminar_operario(request, id):
    """Vista para eliminar un operario."""
    operario = get_object_or_404(Operarios, pk=id)
    if request.method == 'POST':
        try:
            operario.delete()
            messages.success(request, 'Operario eliminado exitosamente.')
            return redirect('gestion:operarios')
        except Exception as e:
            messages.error(request, f'Error al eliminar el operario: {str(e)}')
    return redirect('gestion:operarios')

@login_required
@transaction.atomic
def eliminar_tercero(request, id):
    """Vista para eliminar un tercero."""
    tercero = get_object_or_404(Terceros, pk=id)
    if request.method == 'POST':
        try:
            tercero.delete()
            messages.success(request, 'Tercero eliminado exitosamente.')
            return redirect('gestion:terceros')
        except Exception as e:
            messages.error(request, f'Error al eliminar el tercero: {str(e)}')
    return redirect('gestion:terceros')

@login_required
@transaction.atomic
def eliminar_bodega(request, id):
    """Vista para eliminar una bodega."""
    bodega = get_object_or_404(Bodegas, pk=id)
    if request.method == 'POST':
        try:
            bodega.delete()
            messages.success(request, 'Bodega eliminada exitosamente.')
            return redirect('gestion:bodegas')
        except Exception as e:
            messages.error(request, f'Error al eliminar la bodega: {str(e)}')
    return redirect('gestion:bodegas')

@login_required
def motivos_paro(request):
    """Vista para gestionar los motivos de paro."""
    if request.method == 'POST':
        try:
            nombre = request.POST.get('nombre')
            MotivoParo.objects.create(nombre=nombre)
            messages.success(request, 'Motivo de paro creado exitosamente.')
            return redirect('gestion:motivos_paro')
        except Exception as e:
            messages.error(request, f'Error al crear el motivo de paro: {str(e)}')

    motivos = MotivoParo.objects.all()
    context = {
        'motivos': motivos,
    }
    return render(request, 'gestion/motivos_paro.html', context)

@login_required
@transaction.atomic
def editar_motivo_paro(request, id):
    """Vista para editar un motivo de paro."""
    motivo = get_object_or_404(MotivoParo, pk=id)
    if request.method == 'POST':
        try:
            motivo.nombre = request.POST.get('nombre')
            motivo.save()
            messages.success(request, 'Motivo de paro actualizado exitosamente.')
            return redirect('gestion:motivos_paro')
        except Exception as e:
            messages.error(request, f'Error al actualizar el motivo de paro: {str(e)}')
    
    context = {
        'motivo': motivo,
    }
    return render(request, 'gestion/motivos_paro.html', context)

@login_required
@transaction.atomic
def eliminar_motivo_paro(request, id):
    """Vista para eliminar un motivo de paro."""
    motivo = get_object_or_404(MotivoParo, pk=id)
    if request.method == 'POST':
        try:
            motivo.delete()
            messages.success(request, 'Motivo de paro eliminado exitosamente.')
            return redirect('gestion:motivos_paro')
        except Exception as e:
            messages.error(request, f'Error al eliminar el motivo de paro: {str(e)}')
    return redirect('gestion:motivos_paro')

# --- Funciones de importación de Excel ---

@login_required
def generar_plantilla_materiales(request):
    """Genera una plantilla Excel para importar materiales."""
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plantilla Materiales"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ['Nombre', 'Tipo', 'Descripción']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Datos de ejemplo
    ejemplos = [
        ['Polietileno de Alta Densidad', 'MP', 'Material plástico para inyección'],
        ['Producto Molido Blanco', 'PI', 'Material procesado en molino'],
        ['Bolsas Comerciales', 'PT', 'Producto terminado para venta'],
        ['Colorante Azul', 'IN', 'Aditivo para colorear materiales']
    ]
    
    for row, ejemplo in enumerate(ejemplos, 2):
        for col, valor in enumerate(ejemplo, 1):
            cell = ws.cell(row=row, column=col, value=valor)
            cell.border = thin_border
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 50
    
    # Crear hoja de instrucciones
    ws_inst = wb.create_sheet("Instrucciones")
    
    # Instrucciones
    instrucciones = [
        "INSTRUCCIONES PARA IMPORTAR MATERIALES",
        "",
        "1. Complete la información en la hoja 'Plantilla Materiales'",
        "",
        "2. Campos obligatorios:",
        "   - Nombre: Nombre único del material",
        "   - Tipo: Debe ser uno de los siguientes valores:",
        "     * MP = Materia Prima",
        "     * PI = Producto Intermedio", 
        "     * PT = Producto Terminado",
        "     * IN = Insumo",
        "",
        "3. Campos opcionales:",
        "   - Descripción: Información adicional del material",
        "",
        "4. Notas importantes:",
        "   - No modifique los encabezados de las columnas",
        "   - Los nombres de materiales deben ser únicos",
        "   - Use exactamente los códigos de tipo especificados",
        "   - Puede eliminar las filas de ejemplo antes de importar",
        "",
        "5. Una vez completado, guarde el archivo y úselo en",
        "   la función 'Importar desde Excel' del sistema."
    ]
    
    for row, instruccion in enumerate(instrucciones, 1):
        cell = ws_inst.cell(row=row, column=1, value=instruccion)
        if row == 1:
            cell.font = Font(bold=True, size=14)
        elif instruccion.startswith(("1.", "2.", "3.", "4.", "5.")):
            cell.font = Font(bold=True)
    
    ws_inst.column_dimensions['A'].width = 60
    
    # Generar respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="plantilla_materiales.xlsx"'
    wb.save(response)
    
    return response

@login_required
@transaction.atomic
def importar_materiales_excel(request):
    """Procesa la importación de materiales desde un archivo Excel."""
    if request.method == 'POST' and request.FILES.get('archivo_excel'):
        try:
            archivo = request.FILES['archivo_excel']
            
            # Validar extensión del archivo
            if not archivo.name.endswith(('.xlsx', '.xls')):
                messages.error(request, 'Por favor seleccione un archivo Excel válido (.xlsx o .xls)')
                return redirect('gestion:materiales')
            
            # Leer archivo Excel
            try:
                df = pd.read_excel(archivo)
            except Exception as e:
                messages.error(request, f'Error al leer el archivo Excel: {str(e)}')
                return redirect('gestion:materiales')
            
            # Validar columnas requeridas
            columnas_requeridas = ['Nombre', 'Tipo']
            columnas_faltantes = [col for col in columnas_requeridas if col not in df.columns]
            
            if columnas_faltantes:
                messages.error(request, f'Faltan las siguientes columnas: {", ".join(columnas_faltantes)}')
                return redirect('gestion:materiales')
            
            # Tipos válidos de material
            tipos_validos = dict(Materiales.TIPO_MATERIAL_CHOICES)
            
            materiales_creados = 0
            materiales_omitidos = 0
            errores = []
            
            for index, row in df.iterrows():
                try:
                    # Validar datos de la fila
                    nombre = str(row['Nombre']).strip() if pd.notna(row['Nombre']) else ''
                    tipo = str(row['Tipo']).strip().upper() if pd.notna(row['Tipo']) else ''
                    descripcion = str(row.get('Descripción', '')).strip() if pd.notna(row.get('Descripción', '')) else ''
                    
                    # Validaciones
                    if not nombre:
                        errores.append(f'Fila {index + 2}: El nombre es obligatorio')
                        continue
                    
                    if not tipo:
                        errores.append(f'Fila {index + 2}: El tipo es obligatorio')
                        continue
                    
                    if tipo not in tipos_validos:
                        errores.append(f'Fila {index + 2}: Tipo "{tipo}" no válido. Use: {", ".join(tipos_validos.keys())}')
                        continue
                    
                    # Verificar si el material ya existe
                    if Materiales.objects.filter(nombre__iexact=nombre).exists():
                        materiales_omitidos += 1
                        errores.append(f'Fila {index + 2}: Material "{nombre}" ya existe')
                        continue
                    
                    # Crear material
                    Materiales.objects.create(
                        nombre=nombre,
                        tipo=tipo,
                        descripcion=descripcion
                    )
                    materiales_creados += 1
                    
                except Exception as e:
                    errores.append(f'Fila {index + 2}: Error al procesar - {str(e)}')
            
            # Mensajes de resultado
            if materiales_creados > 0:
                messages.success(request, f'Se importaron {materiales_creados} materiales exitosamente.')
            
            if materiales_omitidos > 0:
                messages.warning(request, f'{materiales_omitidos} materiales fueron omitidos por ya existir.')
            
            if errores:
                error_msg = f'Se encontraron {len(errores)} errores:\n' + '\n'.join(errores[:10])
                if len(errores) > 10:
                    error_msg += f'\n... y {len(errores) - 10} errores más.'
                messages.error(request, error_msg)
            
        except Exception as e:
            messages.error(request, f'Error general al importar archivo: {str(e)}')
    
    return redirect('gestion:materiales')

# --- Funciones de importación de Excel para Operarios ---

@login_required
def generar_plantilla_operarios(request):
    """Genera una plantilla Excel para importar operarios."""
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plantilla Operarios"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ['Código', 'Nombre Completo', 'Activo']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Datos de ejemplo
    ejemplos = [
        ['OP001', 'Juan Carlos Pérez García', 'SI'],
        ['OP002', 'María Elena Rodríguez López', 'SI'],
        ['OP003', 'Carlos Alberto Martínez Sánchez', 'NO'],
        ['OP004', 'Ana Isabel González Fernández', 'SI']
    ]
    
    for row, ejemplo in enumerate(ejemplos, 2):
        for col, valor in enumerate(ejemplo, 1):
            cell = ws.cell(row=row, column=col, value=valor)
            cell.border = thin_border
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 15
    
    # Crear hoja de instrucciones
    ws_inst = wb.create_sheet("Instrucciones")
    
    # Instrucciones
    instrucciones = [
        "INSTRUCCIONES PARA IMPORTAR OPERARIOS",
        "",
        "1. Complete la información en la hoja 'Plantilla Operarios'",
        "",
        "2. Campos obligatorios:",
        "   - Código: Código único del operario (ej: OP001)",
        "   - Nombre Completo: Nombre completo del operario",
        "",
        "3. Campos opcionales:",
        "   - Activo: SI o NO (por defecto será SI si se deja vacío)",
        "",
        "4. Notas importantes:",
        "   - No modifique los encabezados de las columnas",
        "   - Los códigos de operarios deben ser únicos",
        "   - Para el campo Activo use: SI, NO, 1, 0, True, False",
        "   - Puede eliminar las filas de ejemplo antes de importar",
        "",
        "5. Una vez completado, guarde el archivo y úselo en",
        "   la función 'Importar desde Excel' del sistema."
    ]
    
    for row, instruccion in enumerate(instrucciones, 1):
        cell = ws_inst.cell(row=row, column=1, value=instruccion)
        if row == 1:
            cell.font = Font(bold=True, size=14)
        elif instruccion.startswith(("1.", "2.", "3.", "4.", "5.")):
            cell.font = Font(bold=True)
    
    ws_inst.column_dimensions['A'].width = 60
    
    # Generar respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="plantilla_operarios.xlsx"'
    wb.save(response)
    
    return response

@login_required
@transaction.atomic
def importar_operarios_excel(request):
    """Procesa la importación de operarios desde un archivo Excel."""
    if request.method == 'POST' and request.FILES.get('archivo_excel'):
        try:
            archivo = request.FILES['archivo_excel']
            
            # Validar extensión del archivo
            if not archivo.name.endswith(('.xlsx', '.xls')):
                messages.error(request, 'Por favor seleccione un archivo Excel válido (.xlsx o .xls)')
                return redirect('gestion:operarios')
            
            # Leer archivo Excel
            try:
                df = pd.read_excel(archivo)
            except Exception as e:
                messages.error(request, f'Error al leer el archivo Excel: {str(e)}')
                return redirect('gestion:operarios')
            
            # Validar columnas requeridas
            columnas_requeridas = ['Código', 'Nombre Completo']
            columnas_faltantes = [col for col in columnas_requeridas if col not in df.columns]
            
            if columnas_faltantes:
                messages.error(request, f'Faltan las siguientes columnas: {", ".join(columnas_faltantes)}')
                return redirect('gestion:operarios')
            
            operarios_creados = 0
            operarios_omitidos = 0
            errores = []
            
            for index, row in df.iterrows():
                try:
                    # Validar datos de la fila
                    codigo = str(row['Código']).strip() if pd.notna(row['Código']) else ''
                    nombre_completo = str(row['Nombre Completo']).strip() if pd.notna(row['Nombre Completo']) else ''
                    activo_str = str(row.get('Activo', 'SI')).strip().upper() if pd.notna(row.get('Activo', 'SI')) else 'SI'
                    
                    # Convertir activo a booleano
                    activo = activo_str in ['SI', 'SÍ', 'TRUE', '1', 'VERDADERO', 'YES', 'Y']
                    
                    # Validaciones
                    if not codigo:
                        errores.append(f'Fila {index + 2}: El código es obligatorio')
                        continue
                    
                    if not nombre_completo:
                        errores.append(f'Fila {index + 2}: El nombre completo es obligatorio')
                        continue
                    
                    # Verificar si el operario ya existe
                    if Operarios.objects.filter(codigo__iexact=codigo).exists():
                        operarios_omitidos += 1
                        errores.append(f'Fila {index + 2}: Operario con código "{codigo}" ya existe')
                        continue
                    
                    # Crear operario
                    Operarios.objects.create(
                        codigo=codigo,
                        nombre_completo=nombre_completo,
                        activo=activo
                    )
                    operarios_creados += 1
                    
                except Exception as e:
                    errores.append(f'Fila {index + 2}: Error al procesar - {str(e)}')
            
            # Mensajes de resultado
            if operarios_creados > 0:
                messages.success(request, f'Se importaron {operarios_creados} operarios exitosamente.')
            
            if operarios_omitidos > 0:
                messages.warning(request, f'{operarios_omitidos} operarios fueron omitidos por ya existir.')
            
            if errores:
                error_msg = f'Se encontraron {len(errores)} errores:\n' + '\n'.join(errores[:10])
                if len(errores) > 10:
                    error_msg += f'\n... y {len(errores) - 10} errores más.'
                messages.error(request, error_msg)
            
        except Exception as e:
            messages.error(request, f'Error general al importar archivo: {str(e)}')
    
    return redirect('gestion:operarios')

# --- Funciones de importación de Excel para Terceros ---

@login_required
def generar_plantilla_terceros(request):
    """Genera una plantilla Excel para importar terceros."""
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plantilla Terceros"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ['Nombre', 'Tipo', 'Identificación', 'Dirección', 'Teléfono', 'Email']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Datos de ejemplo
    ejemplos = [
        ['Proveedor Plásticos SA', 'Proveedor', '900123456-7', 'Calle 123 #45-67', '3001234567', 'ventas@proveedorplasticos.com'],
        ['Cliente Industrial LTDA', 'Cliente', '800987654-3', 'Carrera 45 #12-34', '3009876543', 'compras@clienteindustrial.com'],
        ['Transportes Rápidos SAS', 'Proveedor', '900555444-2', 'Avenida 68 #23-45', '3005554444', 'info@transportesrapidos.com'],
        ['Empresa Manufacturera', 'Cliente', '800111222-1', 'Zona Industrial Km 15', '3001112222', 'gerencia@manufactureraem.com']
    ]
    
    for row, ejemplo in enumerate(ejemplos, 2):
        for col, valor in enumerate(ejemplo, 1):
            cell = ws.cell(row=row, column=col, value=valor)
            cell.border = thin_border
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 35
    
    # Crear hoja de instrucciones
    ws_inst = wb.create_sheet("Instrucciones")
    
    # Instrucciones
    instrucciones = [
        "INSTRUCCIONES PARA IMPORTAR TERCEROS",
        "",
        "1. Complete la información en la hoja 'Plantilla Terceros'",
        "",
        "2. Campos obligatorios:",
        "   - Nombre: Nombre de la empresa o persona",
        "   - Tipo: Debe ser uno de los siguientes valores:",
        "     * Proveedor",
        "     * Cliente",
        "     * Empleado",
        "     * Otro",
        "",
        "3. Campos opcionales:",
        "   - Identificación: NIT, CC, CE, etc.",
        "   - Dirección: Dirección física",
        "   - Teléfono: Número de contacto",
        "   - Email: Correo electrónico",
        "",
        "4. Notas importantes:",
        "   - No modifique los encabezados de las columnas",
        "   - Los nombres deben ser únicos",
        "   - Use exactamente los tipos especificados",
        "   - Puede eliminar las filas de ejemplo antes de importar",
        "",
        "5. Una vez completado, guarde el archivo y úselo en",
        "   la función 'Importar desde Excel' del sistema."
    ]
    
    for row, instruccion in enumerate(instrucciones, 1):
        cell = ws_inst.cell(row=row, column=1, value=instruccion)
        if row == 1:
            cell.font = Font(bold=True, size=14)
        elif instruccion.startswith(("1.", "2.", "3.", "4.", "5.")):
            cell.font = Font(bold=True)
    
    ws_inst.column_dimensions['A'].width = 60
    
    # Generar respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="plantilla_terceros.xlsx"'
    wb.save(response)
    
    return response

@login_required
@transaction.atomic
def importar_terceros_excel(request):
    """Procesa la importación de terceros desde un archivo Excel."""
    if request.method == 'POST' and request.FILES.get('archivo_excel'):
        try:
            archivo = request.FILES['archivo_excel']
            
            # Validar extensión del archivo
            if not archivo.name.endswith(('.xlsx', '.xls')):
                messages.error(request, 'Por favor seleccione un archivo Excel válido (.xlsx o .xls)')
                return redirect('gestion:terceros')
            
            # Leer archivo Excel
            try:
                df = pd.read_excel(archivo)
            except Exception as e:
                messages.error(request, f'Error al leer el archivo Excel: {str(e)}')
                return redirect('gestion:terceros')
            
            # Validar columnas requeridas
            columnas_requeridas = ['Nombre', 'Tipo']
            columnas_faltantes = [col for col in columnas_requeridas if col not in df.columns]
            
            if columnas_faltantes:
                messages.error(request, f'Faltan las siguientes columnas: {", ".join(columnas_faltantes)}')
                return redirect('gestion:terceros')
            
            # Tipos válidos de tercero
            tipos_validos = ['Proveedor', 'Cliente', 'Empleado', 'Otro']
            
            terceros_creados = 0
            terceros_omitidos = 0
            errores = []
            
            for index, row in df.iterrows():
                try:
                    # Validar datos de la fila
                    nombre = str(row['Nombre']).strip() if pd.notna(row['Nombre']) else ''
                    tipo = str(row['Tipo']).strip() if pd.notna(row['Tipo']) else ''
                    identificacion = str(row.get('Identificación', '')).strip() if pd.notna(row.get('Identificación', '')) else ''
                    direccion = str(row.get('Dirección', '')).strip() if pd.notna(row.get('Dirección', '')) else ''
                    telefono = str(row.get('Teléfono', '')).strip() if pd.notna(row.get('Teléfono', '')) else ''
                    email = str(row.get('Email', '')).strip() if pd.notna(row.get('Email', '')) else ''
                    
                    # Validaciones
                    if not nombre:
                        errores.append(f'Fila {index + 2}: El nombre es obligatorio')
                        continue
                    
                    if not tipo:
                        errores.append(f'Fila {index + 2}: El tipo es obligatorio')
                        continue
                    
                    if tipo not in tipos_validos:
                        errores.append(f'Fila {index + 2}: Tipo "{tipo}" no válido. Use: {", ".join(tipos_validos)}')
                        continue
                    
                    # Verificar si el tercero ya existe
                    if Terceros.objects.filter(nombre__iexact=nombre).exists():
                        terceros_omitidos += 1
                        errores.append(f'Fila {index + 2}: Tercero "{nombre}" ya existe')
                        continue
                    
                    # Crear tercero
                    Terceros.objects.create(
                        nombre=nombre,
                        tipo=tipo,
                        identificacion=identificacion,
                        direccion=direccion,
                        telefono=telefono,
                        email=email,
                        activo=True
                    )
                    terceros_creados += 1
                    
                except Exception as e:
                    errores.append(f'Fila {index + 2}: Error al procesar - {str(e)}')
            
            # Mensajes de resultado
            if terceros_creados > 0:
                messages.success(request, f'Se importaron {terceros_creados} terceros exitosamente.')
            
            if terceros_omitidos > 0:
                messages.warning(request, f'{terceros_omitidos} terceros fueron omitidos por ya existir.')
            
            if errores:
                error_msg = f'Se encontraron {len(errores)} errores:\n' + '\n'.join(errores[:10])
                if len(errores) > 10:
                    error_msg += f'\n... y {len(errores) - 10} errores más.'
                messages.error(request, error_msg)
            
        except Exception as e:
            messages.error(request, f'Error general al importar archivo: {str(e)}')
    
    return redirect('gestion:terceros')
